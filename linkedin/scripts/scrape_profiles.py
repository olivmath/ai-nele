"""Scrape LinkedIn profiles via Playwright (uses existing Chrome session)."""
import json
import os
import re
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

OUTPUT = Path("profiles_results.json")
SPREADSHEET = Path("Floripa_Web3_Tracking_Ligacoes.xlsx")

JS_EXTRACT = """
(() => {
  const t = document.title.replace(/\\s*\\|?\\s*LinkedIn\\s*$/i, '').trim();
  let nm = t, co = '';
  if (t.includes(' - ')) {
    const p = t.split(' - ');
    nm = p[0].trim();
    co = p.slice(1).join(' - ').trim();
  }
  const sec = document.querySelector('main section');
  const txt = sec?.innerText || '';
  const lines = txt.split('\\n').map(x => x.trim()).filter(x => x.length > 2);
  let hl = '?', loc = '?';
  const skip = /seguidores|conex|Dados de contato|Informações|Enviar mensag|^Conectar$|^Seguir$|Pendente|Foto|notifica|Gerenciar|^Mais$|^Sobre$|Experimente/;
  for (const x of lines) {
    if (skip.test(x)) continue;
    if (x.startsWith('·') || x.startsWith('•')) continue;
    if (x.toUpperCase() === nm.toUpperCase()) continue;
    if (hl === '?') { hl = x; continue; }
    if (/Brasil|Brazil|United|Portugal|Argentina|Canada|Germany|France|Florianópolis|São Paulo|Rio de|Santa Catarina|Curitiba|Paraná|Minas|Goiânia|Brasília|Recife|Salvador|Belo Horizonte|Porto Alegre|Joinville|Blumenau|Itajaí|São José|Palhoça|Biguaçu/.test(x)) {
      loc = x.replace(/\\s*(Informações|Dados)\\s*de\\s*contato/g, '').trim();
      break;
    }
  }
  return JSON.stringify({name: nm, company: co, headline: hl, location: loc});
})()
"""


def extract_public_id(url: str) -> str | None:
    if not url or url.strip() == "-" or url.strip() == "None":
        return None
    url = url.strip().rstrip("/")
    if "/in/" in url:
        pid = url.split("/in/")[-1].split("?")[0].strip()
        if pid and " " not in pid:
            return pid
    return None


def load_queue() -> list[dict]:
    wb = openpyxl.load_workbook(SPREADSHEET)
    ws = wb["Contatos"]
    queue = []
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True)):
        name = row[1]
        linkedin = str(row[10]) if row[10] else "-"
        if not name:
            continue
        pub_id = extract_public_id(linkedin)
        if not pub_id:
            continue
        queue.append({
            "row": i + 4,
            "sheet_name": name,
            "pub_id": pub_id,
            "url": f"https://linkedin.com/in/{pub_id}",
        })
    return queue


def load_existing() -> dict:
    if OUTPUT.exists():
        with open(OUTPUT) as f:
            data = json.load(f)
        return {r["pub_id"]: r for r in data}
    return {}


def save_results(results: dict):
    with open(OUTPUT, "w") as f:
        json.dump(list(results.values()), f, indent=2, ensure_ascii=False)


def main():
    queue = load_queue()
    existing = load_existing()
    todo = [q for q in queue if q["pub_id"] not in existing]
    print(f"Total: {len(queue)} | Already done: {len(existing)} | Todo: {len(todo)}")

    if not todo:
        print("Nothing to do!")
        return

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        context = browser.contexts[0]
        page = context.new_page()

        for i, item in enumerate(todo):
            pub_id = item["pub_id"]
            print(f"[{i+1}/{len(todo)}] {item['sheet_name']} ({pub_id})")

            try:
                page.goto(item["url"], wait_until="domcontentloaded", timeout=15000)
                page.wait_for_timeout(3000)

                result_json = page.evaluate(JS_EXTRACT)
                data = json.loads(result_json)
                data["pub_id"] = pub_id
                data["sheet_name"] = item["sheet_name"]
                data["row"] = item["row"]

                if data.get("headline") == "?" and data.get("name") == "":
                    data["error"] = "not_found"
                    print(f"  -> NOT FOUND")
                else:
                    print(f"  -> {data.get('headline', '?')[:60]} | {data.get('location', '?')}")

                existing[pub_id] = data
                save_results(existing)

            except Exception as e:
                print(f"  -> ERROR: {e}")
                existing[pub_id] = {
                    "pub_id": pub_id,
                    "sheet_name": item["sheet_name"],
                    "row": item["row"],
                    "error": str(e),
                }
                save_results(existing)

            time.sleep(2)

        page.close()
        browser.close()

    print(f"\nDone! {len(existing)} profiles in {OUTPUT}")


if __name__ == "__main__":
    main()
