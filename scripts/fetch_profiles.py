"""Fetch LinkedIn profiles for lead qualification."""
import json
import os
import time

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv()

COOKIE = os.environ["LINKEDIN_SECONDARY_COOKIE"]
JSESSIONID = os.environ["LINKEDIN_SECONDARY_JSESSIONID"]
BASE_URL = "https://www.linkedin.com/voyager/api/identity/dash/profiles"
OUTPUT_FILE = "profiles_data.json"


def make_session():
    s = requests.Session()
    s.max_redirects = 0
    s.cookies.set("li_at", COOKIE, domain=".linkedin.com", path="/")
    s.cookies.set("JSESSIONID", f'"{JSESSIONID}"', domain=".linkedin.com", path="/")
    s.headers.update({
        "csrf-token": f'"{JSESSIONID}"',
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/128.0.0.0 Safari/537.36"
        ),
        "Accept": "application/vnd.linkedin.normalized+json+2.1",
        "x-li-lang": "en_US",
        "x-restli-protocol-version": "2.0.0",
    })
    return s


def fetch_profile(public_id: str) -> dict | None:
    s = make_session()
    url = f"{BASE_URL}?q=memberIdentity&memberIdentity={public_id}"
    try:
        res = s.get(url, allow_redirects=False)
        if res.status_code == 200:
            data = res.json()
            elems = data.get("elements", [])
            if elems:
                e = elems[0]
                return {
                    "public_id": public_id,
                    "firstName": e.get("firstName"),
                    "lastName": e.get("lastName"),
                    "headline": e.get("headline"),
                    "location": e.get("location"),
                    "geoLocation": e.get("geoLocation"),
                }
        print(f"  [{res.status_code}] {public_id}")
        return None
    except Exception as ex:
        print(f"  [ERROR] {public_id}: {ex}")
        return None


def extract_public_id(linkedin_url: str) -> str | None:
    if not linkedin_url or linkedin_url.strip() == "-":
        return None
    url = linkedin_url.strip().rstrip("/")
    if "/in/" in url:
        return url.split("/in/")[-1].split("?")[0].strip()
    return None


def main():
    wb = openpyxl.load_workbook("Floripa_Web3_Tracking_Ligacoes.xlsx")
    ws = wb["Contatos"]

    existing = {}
    if os.path.exists(OUTPUT_FILE):
        with open(OUTPUT_FILE) as f:
            for item in json.load(f):
                existing[item["public_id"]] = item

    contacts = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        name = row[1]
        linkedin = row[10] if len(row) > 10 else None
        if not name:
            continue
        pub_id = extract_public_id(str(linkedin)) if linkedin else None
        contacts.append({"name": name, "linkedin": linkedin, "public_id": pub_id})

    results = list(existing.values())
    total = len([c for c in contacts if c["public_id"] and c["public_id"] not in existing])
    done = 0

    for c in contacts:
        pid = c["public_id"]
        if not pid or pid in existing:
            continue

        done += 1
        print(f"[{done}/{total}] {c['name']} ({pid})")
        profile = fetch_profile(pid)
        if profile:
            profile["sheet_name"] = c["name"]
            results.append(profile)
            existing[pid] = profile
        else:
            results.append({
                "public_id": pid,
                "sheet_name": c["name"],
                "headline": None,
                "location": None,
                "error": True,
            })

        with open(OUTPUT_FILE, "w") as f:
            json.dump(results, f, indent=2, ensure_ascii=False)

        time.sleep(3)

    print(f"\nDone! {len(results)} profiles saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
